#!/usr/bin/env python3
"""Re-export AI architecture.xlsx into app/src/data/*.json."""

from __future__ import annotations

import json
import re
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = ROOT / 'AI architecture.xlsx'
OUT = ROOT / 'app' / 'src' / 'data'


def clean(value: object) -> object | None:
    if value is None:
        return None
    if isinstance(value, float) and value == int(value):
        return int(value)
    if isinstance(value, str):
        text = value.strip()
        return text if text else None
    return value


def to_key(header: object, index: int) -> str:
    if not header:
        return f'col{index}'
    key = re.sub(r'[^a-zA-Z0-9]+', ' ', str(header)).strip().title().replace(' ', '')
    if not key:
        return f'col{index}'
    return key[0].lower() + key[1:]


def sheet_rows(workbook: openpyxl.Workbook, name: str, header_row: int = 4) -> list[dict]:
    worksheet = workbook[name]
    rows = list(worksheet.iter_rows(values_only=True))
    headers = [to_key(header, index) for index, header in enumerate(rows[header_row - 1])]
    data: list[dict] = []
    for row in rows[header_row:]:
        if not any(row):
            continue
        item = {
            key: cleaned
            for key, value in zip(headers, row)
            if (cleaned := clean(value)) is not None
        }
        if item:
            data.append(item)
    return data


def main() -> None:
    workbook = openpyxl.load_workbook(WORKBOOK, data_only=True)
    OUT.mkdir(parents=True, exist_ok=True)

    (OUT / 'architectureMap.json').write_text(
        json.dumps(sheet_rows(workbook, 'Architecture Map'), indent=2, ensure_ascii=False) + '\n'
    )
    (OUT / 'quickPicks.json').write_text(
        json.dumps(sheet_rows(workbook, 'Quick Picks'), indent=2, ensure_ascii=False) + '\n'
    )
    (OUT / 'providerGuide.json').write_text(
        json.dumps(sheet_rows(workbook, 'Provider Guide'), indent=2, ensure_ascii=False) + '\n'
    )
    (OUT / 'databaseDeepDive.json').write_text(
        json.dumps(sheet_rows(workbook, 'Database Deep Dive'), indent=2, ensure_ascii=False) + '\n'
    )
    (OUT / 'aiPlatform.json').write_text(
        json.dumps(sheet_rows(workbook, 'AI Platform Components'), indent=2, ensure_ascii=False)
        + '\n'
    )
    (OUT / 'aiGovernance.json').write_text(
        json.dumps(sheet_rows(workbook, 'AI Governance Controls'), indent=2, ensure_ascii=False)
        + '\n'
    )

    matrix = []
    for row in sheet_rows(workbook, 'Service Decision Matrix'):
        matrix.append(
            {
                'layer': row.get('architectureLayer'),
                'domain': row.get('domain'),
                'subcategory': row.get('subcategory'),
                'decisionLevel': row.get('decisionLevel'),
                'capability': row.get('genericCapability'),
                'purpose': row.get('plainEnglishPurpose'),
                'useWhen': row.get('useWhen'),
                'question': row.get('keyDecisionQuestion'),
                'tradeoff': row.get('keyTradeOff'),
                'ease': row.get('ease'),
                'timeToValue': row.get('timeToFirstValue'),
                'lockIn': row.get('lockInRisk'),
                'aws': row.get('awsAlternativeS'),
                'azure': row.get('azureAlternativeS'),
                'gcp': row.get('googleCloudAlternativeS'),
                'awsScore': row.get('awsScore'),
                'azureScore': row.get('azureScore'),
                'gcpScore': row.get('gcpScore'),
                'bestFit': row.get('defaultBestFit'),
                'recommendation': row.get('recommendation'),
                'equivalence': row.get('equivalence'),
            }
        )
    (OUT / 'serviceMatrix.json').write_text(
        json.dumps([{key: value for key, value in item.items() if value is not None} for item in matrix], indent=2, ensure_ascii=False)
        + '\n'
    )

    ratings_sheet = workbook['Provider Ratings']
    rating_rows = list(ratings_sheet.iter_rows(values_only=True))
    ratings = []
    for row in rating_rows[4:]:
        if not any(row):
            continue
        ratings.append(
            {
                key: value
                for key, value in {
                    'capability': row[0],
                    'layer': row[1],
                    'subcategory': row[2],
                    'decisionLevel': row[3],
                    'question': row[4],
                    'aws': row[5],
                    'azure': row[6],
                    'gcp': row[7],
                    'awsDepth': row[8],
                    'awsEase': row[9],
                    'awsEnterprise': row[10],
                    'awsDataAi': row[11],
                    'awsPortability': row[12],
                    'awsOperations': row[13],
                    'awsCost': row[14],
                    'azureDepth': row[15],
                    'azureEase': row[16],
                    'azureEnterprise': row[17],
                    'azureDataAi': row[18],
                    'azurePortability': row[19],
                    'azureOperations': row[20],
                    'azureCost': row[21],
                    'gcpDepth': row[22],
                    'gcpEase': row[23],
                    'gcpEnterprise': row[24],
                    'gcpDataAi': row[25],
                    'gcpPortability': row[26],
                    'gcpOperations': row[27],
                    'gcpCost': row[28],
                    'awsScore': row[29],
                    'azureScore': row[30],
                    'gcpScore': row[31],
                    'bestFit': row[32],
                    'recommendation': row[33] if len(row) > 33 else None,
                }.items()
                if value is not None and value != ''
            }
        )
    (OUT / 'providerRatings.json').write_text(
        json.dumps(ratings, indent=2, ensure_ascii=False) + '\n'
    )

    start = workbook['START HERE']
    steps = []
    for row in start.iter_rows(min_row=5, max_row=9, values_only=True):
        if row[0]:
            steps.append(
                {
                    'step': clean(row[0]),
                    'title': clean(row[1]),
                    'action': clean(row[2]),
                    'guardrail': clean(row[3]),
                }
            )
    guides = []
    for row in start.iter_rows(min_row=13, max_row=21, values_only=True):
        if row[0]:
            guides.append(
                {
                    'sheet': clean(row[0]),
                    'what': clean(row[1]),
                    'when': clean(row[2]),
                }
            )
    meta = {
        'title': 'Cross-Cloud Service Selection Workbook',
        'subtitle': 'A solution-engineering decision system for AWS, Microsoft Azure and Google Cloud.',
        'steps': steps,
        'guides': guides,
    }
    (OUT / 'meta.json').write_text(json.dumps(meta, indent=2, ensure_ascii=False) + '\n')
    print(f'Exported workbook sheets into {OUT}')


if __name__ == '__main__':
    main()
